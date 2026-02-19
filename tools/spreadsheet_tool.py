"""Spreadsheet Tool - Membuat, membaca, menulis, dan memanipulasi file CSV/Excel."""

import csv
import io
import json
import logging
import os
import time
from typing import Optional

logger = logging.getLogger(__name__)


class SpreadsheetTool:
    def __init__(self, output_dir: str = "data/spreadsheets"):
        self.output_dir = output_dir
        self.spreadsheets: dict[str, dict] = {}
        os.makedirs(output_dir, exist_ok=True)

    async def execute(self, plan: dict) -> str:
        action = plan.get("action", plan.get("intent", ""))
        params = plan.get("params", plan)

        actions = {
            "create": self._handle_create,
            "read": self._handle_read,
            "write": self._handle_write,
            "add_row": self._handle_add_row,
            "add_column": self._handle_add_column,
            "update_cell": self._handle_update_cell,
            "delete_row": self._handle_delete_row,
            "delete_column": self._handle_delete_column,
            "filter": self._handle_filter,
            "sort": self._handle_sort,
            "stats": self._handle_stats,
            "merge": self._handle_merge,
            "export": self._handle_export,
            "search": self._handle_search,
            "pivot": self._handle_pivot,
            "formula": self._handle_formula,
        }

        handler = actions.get(action)
        if handler:
            result = handler(params)
            return json.dumps(result, ensure_ascii=False, default=str)

        return f"Spreadsheet tool siap. Aksi: {', '.join(actions.keys())}"

    def create_spreadsheet(self, name: str, headers: list[str],
                           data: Optional[list[list]] = None,
                           filename: Optional[str] = None) -> dict:
        safe_name = name.replace(" ", "_").lower()
        fname = filename or f"{safe_name}_{int(time.time())}.csv"
        file_path = os.path.join(self.output_dir, fname)

        rows = data or []

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                padded = row + [""] * (len(headers) - len(row)) if len(row) < len(headers) else row[:len(headers)]
                writer.writerow(padded)

        info = {
            "name": name,
            "file_path": file_path,
            "headers": headers,
            "row_count": len(rows),
            "created_at": time.time(),
        }
        self.spreadsheets[safe_name] = info
        logger.info(f"Spreadsheet dibuat: {name} ({len(rows)} baris)")

        return {"success": True, "spreadsheet": info}

    def read_spreadsheet(self, file_path: str, limit: Optional[int] = None,
                         offset: int = 0) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}

        try:
            ext = os.path.splitext(file_path)[1].lower()

            if ext in (".xlsx", ".xls"):
                return self._read_excel(file_path, limit, offset)

            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                all_rows = list(reader)

            if not all_rows:
                return {"success": True, "headers": [], "data": [], "total_rows": 0}

            headers = all_rows[0]
            data_rows = all_rows[1:]
            total = len(data_rows)

            if offset > 0:
                data_rows = data_rows[offset:]
            if limit:
                data_rows = data_rows[:limit]

            return {
                "success": True,
                "headers": headers,
                "data": data_rows,
                "total_rows": total,
                "offset": offset,
                "returned": len(data_rows),
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _read_excel(self, file_path: str, limit: Optional[int], offset: int) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not all_rows:
                return {"success": True, "headers": [], "data": [], "total_rows": 0}

            headers = [str(h) if h else "" for h in all_rows[0]]
            data_rows = [[str(c) if c is not None else "" for c in row] for row in all_rows[1:]]
            total = len(data_rows)

            if offset > 0:
                data_rows = data_rows[offset:]
            if limit:
                data_rows = data_rows[:limit]

            return {
                "success": True,
                "headers": headers,
                "data": data_rows,
                "total_rows": total,
                "offset": offset,
                "returned": len(data_rows),
            }
        except ImportError:
            return {"success": False, "error": "openpyxl tidak terinstal. Jalankan: pip install openpyxl"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def write_csv(self, file_path: str, headers: list[str], data: list[list]) -> dict:
        try:
            os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in data:
                    writer.writerow(row)
            return {"success": True, "file_path": file_path, "rows_written": len(data)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def write_excel(self, file_path: str, headers: list[str], data: list[list],
                    sheet_name: str = "Sheet1") -> dict:
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

            for row_idx, row in enumerate(data, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

            wb.save(file_path)
            return {"success": True, "file_path": file_path, "rows_written": len(data)}
        except ImportError:
            return {"success": False, "error": "openpyxl tidak terinstal"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def add_rows(self, file_path: str, rows: list[list]) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            with open(file_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in rows:
                    writer.writerow(row)
            return {"success": True, "rows_added": len(rows)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def add_column(self, file_path: str, column_name: str,
                   default_value: str = "", values: Optional[list] = None) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"] + [column_name]
            data = read_result["data"]

            for i, row in enumerate(data):
                val = values[i] if values and i < len(values) else default_value
                row.append(str(val))

            return self.write_csv(file_path, headers, data)
        except Exception as e:
            return {"success": False, "error": str(e)}

    def update_cell(self, file_path: str, row_index: int, col_index: int, value: str) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            data = read_result["data"]
            if row_index < 0 or row_index >= len(data):
                return {"success": False, "error": f"Baris {row_index} di luar jangkauan (0-{len(data)-1})"}
            if col_index < 0 or col_index >= len(read_result["headers"]):
                return {"success": False, "error": f"Kolom {col_index} di luar jangkauan"}

            while len(data[row_index]) <= col_index:
                data[row_index].append("")
            data[row_index][col_index] = value

            return self.write_csv(file_path, read_result["headers"], data)
        except Exception as e:
            return {"success": False, "error": str(e)}

    def delete_rows(self, file_path: str, row_indices: list[int]) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            data = [row for i, row in enumerate(read_result["data"]) if i not in row_indices]
            deleted = len(read_result["data"]) - len(data)
            result = self.write_csv(file_path, read_result["headers"], data)
            result["rows_deleted"] = deleted
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    def delete_column(self, file_path: str, col_index: int) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = [h for i, h in enumerate(read_result["headers"]) if i != col_index]
            data = [[c for i, c in enumerate(row) if i != col_index] for row in read_result["data"]]

            return self.write_csv(file_path, headers, data)
        except Exception as e:
            return {"success": False, "error": str(e)}

    def filter_data(self, file_path: str, column: str, operator: str, value: str) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            if column not in headers:
                return {"success": False, "error": f"Kolom '{column}' tidak ditemukan"}

            col_idx = headers.index(column)
            filtered = []

            for row in read_result["data"]:
                cell_val = row[col_idx] if col_idx < len(row) else ""
                match = False

                if operator == "eq":
                    match = cell_val == value
                elif operator == "ne":
                    match = cell_val != value
                elif operator == "contains":
                    match = value.lower() in cell_val.lower()
                elif operator == "starts_with":
                    match = cell_val.lower().startswith(value.lower())
                elif operator == "ends_with":
                    match = cell_val.lower().endswith(value.lower())
                elif operator in ("gt", "lt", "gte", "lte"):
                    try:
                        num_cell = float(cell_val)
                        num_val = float(value)
                        if operator == "gt":
                            match = num_cell > num_val
                        elif operator == "lt":
                            match = num_cell < num_val
                        elif operator == "gte":
                            match = num_cell >= num_val
                        elif operator == "lte":
                            match = num_cell <= num_val
                    except ValueError:
                        match = False
                elif operator == "empty":
                    match = not cell_val.strip()
                elif operator == "not_empty":
                    match = bool(cell_val.strip())

                if match:
                    filtered.append(row)

            return {
                "success": True,
                "headers": headers,
                "data": filtered,
                "total_rows": len(filtered),
                "filter": {"column": column, "operator": operator, "value": value},
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def sort_data(self, file_path: str, column: str, ascending: bool = True) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            if column not in headers:
                return {"success": False, "error": f"Kolom '{column}' tidak ditemukan"}

            col_idx = headers.index(column)

            def sort_key(row):
                val = row[col_idx] if col_idx < len(row) else ""
                try:
                    return (0, float(val))
                except ValueError:
                    return (1, val.lower())

            sorted_data = sorted(read_result["data"], key=sort_key, reverse=not ascending)
            result = self.write_csv(file_path, headers, sorted_data)
            result["sorted_by"] = column
            result["ascending"] = ascending
            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_statistics(self, file_path: str, column: Optional[str] = None) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            data = read_result["data"]
            stats = {
                "total_rows": len(data),
                "total_columns": len(headers),
                "headers": headers,
            }

            columns_to_analyze = [column] if column and column in headers else headers

            col_stats = {}
            for col_name in columns_to_analyze:
                col_idx = headers.index(col_name)
                values = [row[col_idx] if col_idx < len(row) else "" for row in data]
                non_empty = [v for v in values if v.strip()]

                col_info = {
                    "total": len(values),
                    "non_empty": len(non_empty),
                    "empty": len(values) - len(non_empty),
                    "unique": len(set(non_empty)),
                }

                numeric_vals = []
                for v in non_empty:
                    try:
                        numeric_vals.append(float(v))
                    except ValueError:
                        pass

                if numeric_vals:
                    col_info["numeric_count"] = len(numeric_vals)
                    col_info["min"] = min(numeric_vals)
                    col_info["max"] = max(numeric_vals)
                    col_info["sum"] = sum(numeric_vals)
                    col_info["mean"] = round(sum(numeric_vals) / len(numeric_vals), 4)

                    sorted_vals = sorted(numeric_vals)
                    n = len(sorted_vals)
                    col_info["median"] = sorted_vals[n // 2] if n % 2 else (sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2

                col_stats[col_name] = col_info

            stats["columns"] = col_stats
            return {"success": True, "statistics": stats}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def merge_spreadsheets(self, file_paths: list[str], output_path: str,
                           merge_type: str = "vertical") -> dict:
        try:
            all_headers = []
            all_data = []

            for fp in file_paths:
                result = self.read_spreadsheet(fp)
                if not result["success"]:
                    return {"success": False, "error": f"Gagal membaca {fp}: {result.get('error', '')}"}

                if merge_type == "vertical":
                    if not all_headers:
                        all_headers = result["headers"]
                    all_data.extend(result["data"])
                elif merge_type == "horizontal":
                    if not all_headers:
                        all_headers = result["headers"]
                    else:
                        all_headers.extend(result["headers"])

                    if not all_data:
                        all_data = result["data"]
                    else:
                        for i, row in enumerate(result["data"]):
                            if i < len(all_data):
                                all_data[i].extend(row)
                            else:
                                padding = [""] * (len(all_headers) - len(result["headers"]))
                                all_data.append(padding + row)

            return self.write_csv(output_path, all_headers, all_data)
        except Exception as e:
            return {"success": False, "error": str(e)}

    def search_data(self, file_path: str, query: str, columns: Optional[list[str]] = None) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            search_indices = []
            if columns:
                search_indices = [headers.index(c) for c in columns if c in headers]
            else:
                search_indices = list(range(len(headers)))

            query_lower = query.lower()
            results = []
            for row_idx, row in enumerate(read_result["data"]):
                for col_idx in search_indices:
                    if col_idx < len(row) and query_lower in str(row[col_idx]).lower():
                        results.append({
                            "row_index": row_idx,
                            "column": headers[col_idx],
                            "value": row[col_idx],
                            "row": row,
                        })
                        break

            return {
                "success": True,
                "query": query,
                "results": results,
                "total_matches": len(results),
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def pivot_table(self, file_path: str, row_field: str, col_field: str,
                    value_field: str, agg_func: str = "sum") -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            for field in [row_field, col_field, value_field]:
                if field not in headers:
                    return {"success": False, "error": f"Field '{field}' tidak ditemukan"}

            ri = headers.index(row_field)
            ci = headers.index(col_field)
            vi = headers.index(value_field)

            pivot_data: dict[str, dict[str, list]] = {}
            col_values = set()

            for row in read_result["data"]:
                rv = row[ri] if ri < len(row) else ""
                cv = row[ci] if ci < len(row) else ""
                vv = row[vi] if vi < len(row) else "0"

                col_values.add(cv)
                if rv not in pivot_data:
                    pivot_data[rv] = {}
                if cv not in pivot_data[rv]:
                    pivot_data[rv][cv] = []
                try:
                    pivot_data[rv][cv].append(float(vv))
                except ValueError:
                    pivot_data[rv][cv].append(0)

            sorted_cols = sorted(col_values)
            pivot_headers = [row_field] + sorted_cols
            pivot_rows = []

            for rv in sorted(pivot_data.keys()):
                row = [rv]
                for cv in sorted_cols:
                    vals = pivot_data[rv].get(cv, [0])
                    if agg_func == "sum":
                        row.append(str(sum(vals)))
                    elif agg_func == "avg":
                        row.append(str(round(sum(vals) / len(vals), 2)))
                    elif agg_func == "count":
                        row.append(str(len(vals)))
                    elif agg_func == "min":
                        row.append(str(min(vals)))
                    elif agg_func == "max":
                        row.append(str(max(vals)))
                    else:
                        row.append(str(sum(vals)))
                pivot_rows.append(row)

            return {
                "success": True,
                "headers": pivot_headers,
                "data": pivot_rows,
                "total_rows": len(pivot_rows),
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def apply_formula(self, file_path: str, target_column: str, formula: str,
                      source_columns: list[str]) -> dict:
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File tidak ditemukan: {file_path}"}
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            data = read_result["data"]

            col_indices = {}
            for col in source_columns:
                if col in headers:
                    col_indices[col] = headers.index(col)

            if target_column not in headers:
                headers.append(target_column)
                for row in data:
                    row.append("")
            target_idx = headers.index(target_column)

            supported_formulas = {
                "sum": lambda vals: str(sum(vals)),
                "avg": lambda vals: str(round(sum(vals) / len(vals), 2)) if vals else "0",
                "min": lambda vals: str(min(vals)) if vals else "0",
                "max": lambda vals: str(max(vals)) if vals else "0",
                "count": lambda vals: str(len(vals)),
                "concat": lambda vals: " ".join(str(v) for v in vals),
                "multiply": lambda vals: str(vals[0] * vals[1]) if len(vals) >= 2 else "0",
                "subtract": lambda vals: str(vals[0] - vals[1]) if len(vals) >= 2 else "0",
                "divide": lambda vals: str(round(vals[0] / vals[1], 4)) if len(vals) >= 2 and vals[1] != 0 else "0",
                "percentage": lambda vals: str(round((vals[0] / vals[1]) * 100, 2)) if len(vals) >= 2 and vals[1] != 0 else "0",
            }

            func = supported_formulas.get(formula)
            if not func:
                return {"success": False, "error": f"Formula '{formula}' tidak didukung. Gunakan: {list(supported_formulas.keys())}"}

            for row in data:
                while len(row) <= target_idx:
                    row.append("")

                if formula == "concat":
                    vals = [row[col_indices[c]] if col_indices[c] < len(row) else "" for c in source_columns if c in col_indices]
                    row[target_idx] = func(vals)
                else:
                    vals = []
                    for c in source_columns:
                        if c in col_indices and col_indices[c] < len(row):
                            try:
                                vals.append(float(row[col_indices[c]]))
                            except ValueError:
                                vals.append(0)
                    try:
                        row[target_idx] = func(vals)
                    except Exception:
                        row[target_idx] = "ERROR"

            return self.write_csv(file_path, headers, data)
        except Exception as e:
            return {"success": False, "error": str(e)}

    def export_to_json(self, file_path: str, output_path: Optional[str] = None) -> dict:
        try:
            read_result = self.read_spreadsheet(file_path)
            if not read_result["success"]:
                return read_result

            headers = read_result["headers"]
            json_data = []
            for row in read_result["data"]:
                obj = {}
                for i, header in enumerate(headers):
                    obj[header] = row[i] if i < len(row) else ""
                json_data.append(obj)

            out = output_path or file_path.rsplit(".", 1)[0] + ".json"
            with open(out, "w", encoding="utf-8") as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False)

            return {"success": True, "file_path": out, "records": len(json_data)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def import_from_json(self, json_path: str, output_path: Optional[str] = None) -> dict:
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                json_data = json.load(f)

            if not isinstance(json_data, list) or not json_data:
                return {"success": False, "error": "JSON harus berupa array of objects"}

            headers = list(json_data[0].keys())
            data = []
            for obj in json_data:
                row = [str(obj.get(h, "")) for h in headers]
                data.append(row)

            out = output_path or json_path.rsplit(".", 1)[0] + ".csv"
            return self.write_csv(out, headers, data)
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _handle_create(self, params: dict) -> dict:
        return self.create_spreadsheet(
            name=params.get("name", "untitled"),
            headers=params.get("headers", []),
            data=params.get("data"),
            filename=params.get("filename"),
        )

    def _handle_read(self, params: dict) -> dict:
        return self.read_spreadsheet(
            file_path=params.get("file_path", ""),
            limit=params.get("limit"),
            offset=params.get("offset", 0),
        )

    def _handle_write(self, params: dict) -> dict:
        fp = params.get("file_path", "")
        ext = os.path.splitext(fp)[1].lower()
        if ext in (".xlsx", ".xls"):
            return self.write_excel(fp, params.get("headers", []), params.get("data", []))
        return self.write_csv(fp, params.get("headers", []), params.get("data", []))

    def _handle_add_row(self, params: dict) -> dict:
        return self.add_rows(params.get("file_path", ""), params.get("rows", []))

    def _handle_add_column(self, params: dict) -> dict:
        return self.add_column(params.get("file_path", ""), params.get("column_name", ""), params.get("default_value", ""))

    def _handle_update_cell(self, params: dict) -> dict:
        return self.update_cell(params.get("file_path", ""), params.get("row", 0), params.get("col", 0), params.get("value", ""))

    def _handle_delete_row(self, params: dict) -> dict:
        return self.delete_rows(params.get("file_path", ""), params.get("indices", []))

    def _handle_delete_column(self, params: dict) -> dict:
        return self.delete_column(params.get("file_path", ""), params.get("col_index", 0))

    def _handle_filter(self, params: dict) -> dict:
        return self.filter_data(params.get("file_path", ""), params.get("column", ""), params.get("operator", "eq"), params.get("value", ""))

    def _handle_sort(self, params: dict) -> dict:
        return self.sort_data(params.get("file_path", ""), params.get("column", ""), params.get("ascending", True))

    def _handle_stats(self, params: dict) -> dict:
        return self.get_statistics(params.get("file_path", ""), params.get("column"))

    def _handle_merge(self, params: dict) -> dict:
        return self.merge_spreadsheets(params.get("file_paths", []), params.get("output_path", ""), params.get("merge_type", "vertical"))

    def _handle_export(self, params: dict) -> dict:
        fmt = params.get("format", "json")
        if fmt == "json":
            return self.export_to_json(params.get("file_path", ""), params.get("output_path"))
        elif fmt in ("xlsx", "excel"):
            read_result = self.read_spreadsheet(params.get("file_path", ""))
            if read_result["success"]:
                out = params.get("output_path") or params.get("file_path", "").rsplit(".", 1)[0] + ".xlsx"
                return self.write_excel(out, read_result["headers"], read_result["data"])
            return read_result
        return {"success": False, "error": f"Format ekspor tidak didukung: {fmt}"}

    def _handle_search(self, params: dict) -> dict:
        return self.search_data(params.get("file_path", ""), params.get("query", ""), params.get("columns"))

    def _handle_pivot(self, params: dict) -> dict:
        return self.pivot_table(params.get("file_path", ""), params.get("row_field", ""), params.get("col_field", ""), params.get("value_field", ""), params.get("agg_func", "sum"))

    def _handle_formula(self, params: dict) -> dict:
        return self.apply_formula(params.get("file_path", ""), params.get("target_column", ""), params.get("formula", ""), params.get("source_columns", []))
